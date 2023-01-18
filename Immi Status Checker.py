from multiprocessing.connection import wait
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

from openpyxl import Workbook

wb = openpyxl.load_workbook("../Immi Credentials.xlsx", data_only=True)
ws = wb["Eduyoung"]

# load usernames

def new_func(ws):
    usernames = []
    for user in ws["D"]:
        if user.value == None:
            continue
        else: usernames.append(user.value)

    passwords = []
    for password in ws["E"]:
        if password.value == None:
            continue
        else: passwords.append(password.value)

    current_status = []
    for status in ws["G"]:
        if status.value == None:
            continue
        else: current_status.append(status.value)
    return usernames,passwords,current_status

usernames, passwords, current_status = new_func(ws)

options = Options()
options.headless = True
# driver = webdriver.Chrome("/Applications/Google\ Chrome.app/Contents/MacOS/chromedriver", options=options)
driver = webdriver.Chrome(options=options)
 
for x in range(1, len(usernames)):

    # Skip if Application Status is already Finalised
    if current_status[x] != "Finalised":
        # Login
        driver.get("https://online.immi.gov.au/ola/app")
        driver.find_element("name", "username").send_keys(usernames[x])
        driver.find_element("name", "password").send_keys(passwords[x])
        driver.find_element("name", "login").send_keys(Keys.ENTER)
        driver.find_element("name", "continue").send_keys(Keys.ENTER)
        
        # Copy Status from Website and Paste to Excel file
        def update_status():
            status_wait = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/section/div/div/div[3]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/p/strong")))
            global status
            status = driver.find_element("xpath", "/html/body/form/section/div/div/div[3]/div/div[2]/div/div/div[2]/div/div[1]/div/div[2]/div/p/strong").text
            status_cell = "G" + str(x + 2)
            ws[status_cell] = status

        update_status()
        
        # Copy Last Update Date and Paste to Excel File 
        
        def last_update_date():
            last_update_date_cell_wait = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/section/div/div/div[3]/div/div[2]/div/div/div[2]/div/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div/div[1]/div/div/div/div/time")))
            global last_update_date_cell
            last_update_date_cell = driver.find_element("xpath", "/html/body/form/section/div/div/div[3]/div/div[2]/div/div/div[2]/div/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div/div[1]/div/div/div/div/time").text
            if last_update_date_cell == "":
                last_update_date()
            
            update_dates = "H" + str(x + 2)
            previous_update = "I" + str(x + 2)
            
            # if there is a new update, current update will be moved to next cell. new update will be at current cell.
            if ws[update_dates] != last_update_date_cell:
                ws[previous_update] = ws[update_dates].value
                print("New Update:", last_update_date_cell)

            ws[update_dates] = last_update_date_cell
        
        last_update_date()
        
        # Print status to terminal for each user
        user = str(usernames[x])
        print("%s is %s" % (user, status))

        # Logout and Save Excel File
        wb.save("../Immi Credentials.xlsx")
        driver.find_element("xpath", "/html/body/form/header/div/div/ol/li[3]/button").click()
        driver.find_element("xpath", "/html/body/header/div/ul/li/div/a").click()

driver.quit()
